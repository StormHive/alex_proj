import argparse
import numpy as np
from datetime import datetime
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from utils.direct_costs import create_direct_costs_spreadsheet
from utils.work_performed import create_work_performed_spreadsheet
from utils.hrs_sheet import create_HRS_Spreadsheet
from utils.budget_sheet import create_budget_spreadsheet
import utils.helper_functions as helper_functions
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from dateutil.relativedelta import relativedelta

def create_combined_workbook(contract_id, last_month_str, work_year, filename, dc_start_year, dc_end_year):
    try:
        last_month = helper_functions.convert_to_datetime(last_month_str)
        pop_data = helper_functions.get_unique_pop_ids()
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
       
        # Getting Contract ID 
        if contract_id is None:
            contract_ids = pop_data['contract_id'].unique()  
        else:
            contract_ids = [contract_id]  

        for contract_id in contract_ids:
            wb = Workbook()
            wb.remove(wb.active)  
            percentages_df = helper_functions.get_contract_percentages(contract_id)
            if percentages_df is None or percentages_df.empty:
                print(f"No percentage data available for contract_id: {contract_id}")
                continue

            # Create HRS and Budget sheets
            for _, row in pop_data[pop_data['contract_id'] == contract_id].iterrows():
                pop_id = row['pop_id']
                headers_data = helper_functions.get_hrs_budget_headers(helper_functions.db_uri, pop_id)
                data3, lookup_month_info, pop_start_date, pop_end_date = helper_functions.get_hrs_and_budget_data(
                    helper_functions.db_uri, pop_id, last_month_str
                )
                
                if data3 is not None and pop_start_date and pop_end_date:
                    pop_start_date = pd.to_datetime(pop_start_date)
                    pop_end_date = pd.to_datetime(pop_end_date)
                    current_date = pop_start_date
                    end_date = pop_end_date

                    # Apply NETWORKDAYS formulas 
                    while current_date <= end_date:
                        year = current_date.year
                        month = current_date.month
                        start_day = current_date.day if current_date == pop_start_date else 1                       
                        last_day_of_month = (current_date + relativedelta(day=31)).day
                        end_day = last_day_of_month if current_date.month != end_date.month or current_date.year != end_date.year else end_date.day
                        
                        formula = helper_functions.holidays_formula(helper_functions.engine, pop_start_date, pop_end_date)
                        current_date += relativedelta(months=1)

                    period_data_list = helper_functions.get_period_number_data(data3)
                    for year in range(dc_start_year, dc_end_year + 1):
                        for month in range(1, 13):
                            month_str = datetime(year, month, 1).strftime("%b-%y")
                            selected_period_number = helper_functions.check_month_period_number(
                                month_str, period_data_list, default_period_number=-1
                            )

                            # Create HRS and Budget sheets
                            if selected_period_number == 0:
                                hrs_sheet_title = "Base HRS"
                                budget_sheet_title = "Base Budget"
                            elif selected_period_number >= 1:
                                hrs_sheet_title = f"OY{selected_period_number} HRS"
                                budget_sheet_title = f"OY{selected_period_number} Budget"
                            else:
                                continue

                            if hrs_sheet_title not in wb.sheetnames:
                                wb = create_HRS_Spreadsheet(
                                    wb, headers_data, data3[data3['PeriodNumber'] == selected_period_number],
                                    lookup_month_info, last_month, selected_period_number,helper_functions.engine
                                )
                                
                            if budget_sheet_title not in wb.sheetnames:
                                wb = create_budget_spreadsheet(
                                    wb, headers_data, data3[data3['PeriodNumber'] == selected_period_number],
                                    lookup_month_info, last_month, selected_period_number, hrs_sheet_title
                                )

            # Create Work Performed and Direct Costs sheets
            for year in range(dc_start_year, dc_end_year + 1):
                data2 = helper_functions.get_work_performed_data(helper_functions.db_uri, contract_id, year)
                if data2 is None or data2.empty:
                    #print(f"No work performed data for the year {year}.")
                    continue
                
                work_performed_sheet_name = f"{year} Work Performed"
                ws_work_performed = wb.create_sheet(title=work_performed_sheet_name)
                for r_idx, row in enumerate(dataframe_to_rows(data2, index=False, header=True), start=1):
                    for c_idx, value in enumerate(row, start=1):
                        ws_work_performed.cell(row=r_idx, column=c_idx, value=value)

                for cell in ws_work_performed[1]:  
                    cell.fill = header_fill  
                    cell.font = header_font

                # Direct Costs sheet
                direct_costs_sheet_name = f"{year} Direct Costs"
                if direct_costs_sheet_name not in wb.sheetnames:
                    ws_direct_costs = wb.create_sheet(title=direct_costs_sheet_name)
                else:
                    ws_direct_costs = wb[direct_costs_sheet_name]

                work_performed_full_months = [cell.value for cell in ws_work_performed['A'] if cell.value]
                actual_months, forecast_months = helper_functions.categorize_months(
                    dc_start_year=year,  
                    dc_end_year=year,   
                    last_month=last_month
                )

                sql_result = helper_functions.check_work_performed_data_exist(helper_functions.engine)

                # Populate Direct Costs
                data1 = helper_functions.get_direct_costs_data(helper_functions.db_uri, contract_id, work_year)
                if data1 is None or data1.empty:
                    print(f"No direct cost data for the year {year}")
                else:
                    data1_with_columns = helper_functions.add_columns(data1)
                    create_direct_costs_spreadsheet(
                        data1_with_columns, ws_direct_costs, wb, None, actual_months, forecast_months, year, last_month_str, contract_id,
                        dc_start_year, dc_end_year, sql_result
                    )

            # Save workbooks 
            contract_filename = f"Contract_{contract_id}_Combined_spreadsheet.xlsx"
            wb.save(contract_filename)
            print(f"\nWorkbook for contract_id {contract_id} successfully saved to {contract_filename}\n")

    except Exception as e:
        print(f"Error during workbook creation: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Excel file from database data.")
    parser.add_argument('--contract_id', type=int, required=False, default=None, help='Contract ID')
    parser.add_argument('--pop_id', type=int, required=False, help='PoP ID')
    parser.add_argument('--work_year', type=str, default=2024, help='Work year')
    parser.add_argument('--last_month', type=str, required=False, default="08/2024", help='Last Month (Optional)')
    parser.add_argument('--filename', type=str, default='Combined_spreadsheet.xlsx', help='Output Excel filename')
    parser.add_argument('--dc_start_year', type=int, default=2023, required=False, help='Start year')
    parser.add_argument('--dc_end_year', type=int, default=2027, required=False, help='End year')

    args = parser.parse_args()
    create_combined_workbook(args.contract_id, args.last_month, int(args.work_year), args.filename, args.dc_start_year, args.dc_end_year)
