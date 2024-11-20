import datetime
import calendar
import pyodbc
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.sql import text
import urllib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from calendar import month_abbr
import argparse
import numpy as np  
from datetime import datetime, timedelta
import openpyxl.utils.cell as cell_utils

driver = 'ODBC Driver 17 for SQL Server'
server = '127.0.0.1,1433'  # Or 'localhost'
database = 'templdb'
username = 'sa'  
password = 'VeryStr0ngP@ssw0rd'

params = urllib.parse.quote_plus(
    f'driver={{{driver}}};'
    f'server={server};'
    f'database={database};'
    f'uid={username};'
    f'pwd={password};'
    'encrypt=no;'
    'trustservercertificate=yes;'
    'connection timeout=30;'
)

db_uri = f"mssql+pyodbc:///?odbc_connect={params}"
engine = create_engine(db_uri)

#check if work performed data is exist for employee
def check_work_performed_data_exist(engine):
    sql_query = """
        select distinct e.employee_id,
        e.LastName + ', ' + e.FirstName  AS EmployeeName,
        DATENAME(month, wp.dateworked) AS [Fiscal Period],
        YEAR(wp.dateworked) AS [Work year],
        FORMAT(wp.dateworked, 'MMM-yy') AS month_match
    from employee e
    LEFT JOIN WorkPerformed wp ON e.employee_id = wp.employee_id
    """
    with engine.connect() as conn:
        sql_result = pd.read_sql_query(sql_query, conn)
    return sql_result

#get unique pop_id
def get_unique_pop_ids():
    query = (
        "select distinct pop_id, contract_id, StartDate, EndDate, PeriodNumberWithinContract AS PeriodNumber from PeriodOfPerformance"
    )
    with engine.connect() as conn:
        result = pd.read_sql(query, conn)
    return result

#get contract percentages
def get_contract_percentages(contract_id):
    try:
        contract_id = int(contract_id)
        query = text("""
        select FringePercentage, OHPercentage, MSPercentage, GAPercentage from ContractLookupInfo
        where Contract_Id = :contract_id
        """)
        with engine.connect() as conn:
            df = pd.read_sql(query, conn, params={'contract_id': contract_id})
        return df if not df.empty else pd.DataFrame()
    except Exception:
        return pd.DataFrame()

#get direct cost data
def get_direct_costs_data(db_uri, contract_id, work_year):
    try:
        engine = create_engine(db_uri)
        query = "EXEC DirectCostSheetEmployeeDetails @contract_id=?, @work_year=?"
        with engine.connect() as conn:
            data = pd.read_sql(query, conn, params=(int(contract_id), int(work_year)))
        return data
    except Exception as e:
        return None

#get work performed data
def get_work_performed_data(db_uri, contract_id, work_year):
    try:
        engine = create_engine(db_uri)
        query = text("""
            EXEC GetWorkPerformedSheetData @contract_id=:contract_id, @work_year=:work_year
        """)
        with engine.connect() as conn:
            data = pd.read_sql(query, conn, params={'contract_id': int(contract_id), 'work_year': int(work_year)})

        #hard codeed columns
        data['GL Account'] = '1'
        data['GL Account Description'] = 'Direct Billable'
        data['Task Code'] = ''
        data['Task Description'] = ''
        data['Cost Element Code'] = '1'
        data['Cost Element Description'] = 'Employee Labor'
        data['Organization Code'] = '1110'
        data['Organization Description'] = '1110 - Client Furnished Site'
        data['Job Burden Pool'] = ''
        data['Raw Cost Total Burden Amount'] = ''
        data['Target Total Burden Amount'] = ''
        data['Billing Rule ID'] = ''

        #ordering columns
        column_order = [
            'Fiscal Period', 'job', 'Job Title', 
            'GL Account', 'GL Account Description', 
            'Task Code', 'Task Description', 
            'Cost Element Code', 'Cost Element Description', 
            'Organization Code', 'Organization Description', 
            'employee_id', 'EmployeeName', 
            'LaborCategoryCode', 'LaborCategoryDescription', 
            'Raw_Cost Hours/QTY', 'Raw Cost Amount', 'Billed_Amount', 'Revenue Amount',
            'Billed Hours/QTY', 'Revenue Hours/QTY', 
            'Job Burden Pool', 'Raw Cost Total Burden Amount', 
            'Target Total Burden Amount', 'Billing Rule ID','Work year','contract_id'
        ]

        data = data[column_order]
        return data  
    except Exception as e:
        return None

#get HRS and Budget data
def get_hrs_and_budget_data(db_uri, pop_id, last_month):
    try:
        engine = create_engine(db_uri)
        with engine.connect() as conn:
            data_query = """
            EXEC GetHrsBudgetSheetData @pop_id=?, @last_month=?"""
            data = pd.read_sql(data_query, conn, params=(pop_id, last_month))
            data['date'] = pd.to_datetime(data['date'])

            lookup_query = "select Year, Month, WorkDaysAvailable, Holidays, VacationFactor from LookupMonthInfo"
            lookup_month_info = pd.read_sql(lookup_query, conn)

            pop_start_date = data['StartDate'].min().strftime("%Y-%m-%d")
            pop_end_date = data['EndDate'].max().strftime("%Y-%m-%d")
       
        return data, lookup_month_info, pop_start_date, pop_end_date
    except Exception as e:
        return None, None, None, None
 
# get holidays
def holidays_formula(engine, pop_start_date, pop_end_date):
    query = """select holiday_date from federalholidays where holiday_date between ? and ?"""
    with engine.connect() as conn:
        holidays = pd.read_sql_query(query, conn, params=(pop_start_date, pop_end_date))
    
    if not holidays.empty:
        all_holidays = [holiday.strftime("%m/%d/%y") for holiday in holidays['holiday_date']]
        holiday_str = ', '.join([f'"{h}"' for h in all_holidays])
    else:
        all_holidays = []
        holiday_str = ""

    def generate_formula(year, month, start_day, end_day):
        return f'=NETWORKDAYS(DATE({year}, {month}, {start_day}), DATE({year}, {month}, {end_day}), {{{holiday_str}}})'
    return generate_formula, all_holidays

#create HRS budget headers
def get_hrs_budget_headers(db_uri, pop_id):
    try:
        engine = create_engine(db_uri)
        with engine.connect() as conn:
            print(pop_id, "POP_Id")
            header_query = """EXEC GetHrsBudgetSheetHeaders @pop_id=?"""
            data = pd.read_sql(header_query, conn, params=(pop_id,))
        return data
    except Exception as e:
        print(f"Error fetching HRS and Budget header data: {e}")
        return None
  
# list contract months
def list_contract_months(pop_start_date, pop_end_date):
    start = datetime.strptime(pop_start_date, "%Y-%m-%d %H:%M:%S") if " " in pop_start_date else datetime.strptime(pop_start_date, "%Y-%m-%d")
    end = datetime.strptime(pop_end_date, "%Y-%m-%d %H:%M:%S") if " " in pop_end_date else datetime.strptime(pop_end_date, "%Y-%m-%d")

    months = []
    while start <= end:
        months.append(start.strftime("%b-%y"))
        if start.month == 12:
            start = start.replace(year=start.year + 1, month=1)
        else:
            start = start.replace(month=start.month + 1)
    end_month_str = end.strftime("%b-%y")
    if months[-1] != end_month_str:
        months.append(end_month_str)
    return months

#range of months
def list_year_months(dc_start_date, dc_end_date, full_year=False):
    start = datetime.strptime(dc_start_date, "%Y-%m-%d %H:%M:%S") if " " in dc_start_date else datetime.strptime(dc_start_date, "%Y-%m-%d")
    end = datetime.strptime(dc_end_date, "%Y-%m-%d %H:%M:%S") if " " in dc_end_date else datetime.strptime(dc_end_date, "%Y-%m-%d")

    months = []
    while start <= end:
        months.append(start.strftime("%b-%y"))
        if start.month == 12:
            start = start.replace(year=start.year + 1, month=1)
        else:
            start = start.replace(month=start.month + 1)
    if full_year:
        end = datetime.strptime(f"{end.year}-12-31", "%Y-%m-%d")
        while start <= end:
            months.append(start.strftime("%b-%y"))
            if start.month == 12:
                start = start.replace(year=start.year + 1, month=1)
            else:
                start = start.replace(month=start.month + 1)
    return months 

#hard coded columns
def add_columns(data):
    if data is None:
        return None
    data['CM%'] = ''
    data['OI%'] = ''
    data['% Direct'] = ''
    data['Employee / Sub'] = 'Employee Labor'
    return data

#sort months
def sort_key(date_str):
    month, year = date_str.split('-')
    month_num = list(month_abbr).index(month)
    return (int(year), month_num)

#compare last month parameter with header months
def convert_to_datetime(month_str):
    try:
        if '/' in month_str:
            month, year = month_str.split('/')
            month_num = int(month)
            year_num = int(year)
            return datetime(year=year_num, month=month_num, day=1)
        elif '-' in month_str:
            month, year = month_str.split('-')
            month_num = datetime.strptime(month, "%b").month
            return datetime(year=int("20" + year), month=month_num, day=1)
        else:
            raise ValueError("Invalid date format")
    except ValueError as ve:
        print(f"Error converting {month_str} to datetime: {ve}")
        return None
    
#get project start and end date
def get_period_number_data(data3):
    period_data = data3[['pop_id', 'StartDate', 'EndDate', 'PeriodNumber']].drop_duplicates()
    period_data_list = list(set([tuple(x) for x in period_data.to_records(index=False)]))
    return period_data_list

#check peroid number for months
def check_month_period_number(month_str, period_data, default_period_number=-1):
    try:
        month_by_period = datetime.strptime(month_str, "%b-%y").date()
    except ValueError as e:
        print(f"Error for month '{month_str}': {e}")
        return default_period_number
    selected_period_number = default_period_number
    for period in period_data:
        pop_id, dc_start_date, dc_end_date, period_number = period

        if isinstance(dc_start_date, datetime):
            dc_start_date = dc_start_date.date()
        if isinstance(dc_end_date, datetime):
            dc_end_date = dc_end_date.date()
        if dc_start_date <= month_by_period <= dc_end_date:
            selected_period_number = period_number
            break  
    return selected_period_number

#create footer
def add_footer_rows(ws, start_row, generated_months):
    total_col_letter = chr(ord('A') + len(generated_months) + 3)
    for i, month in enumerate(generated_months, start=5):
        month_col_letter = chr(ord('A') + i - 1)
        set_cell(ws, f'{month_col_letter}{start_row + 1}', f"=ROUND(SUM({month_col_letter}13:{month_col_letter}{start_row}), 1)")
        ws[f'{month_col_letter}{start_row + 1}'].number_format = '0.0'
        set_cell(ws, f'{total_col_letter}{start_row + 1}', f"=ROUND(SUM({total_col_letter}13:{total_col_letter}{start_row}), 1)")
        ws[f'{total_col_letter}{start_row + 1}'].number_format = '0.0'

def set_cell(ws, cell, value, bold=False, fill=None, align_center=False):
    ws[cell] = value if value is not None else ''
    ws[cell].font = Font(name='Tahoma', size=9, bold=bold)
    if fill:
        ws[cell].fill = fill
    if align_center:
        ws[cell].alignment = Alignment(horizontal='center')

def set_font_for_sheet(ws, font_name='Tahoma', font_size=9):
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name=font_name, size=font_size)

def create_footer_row(ws, group_start_row, end_row, generated_months, last_month_dt, job_name):
    footer_rows = []
    footer_row = end_row + 1
    first_month_col = 'E'

    group_job_name = ws[f'B{group_start_row}'].value.split(":")[0]
    ws[f'B{footer_row}'].value = f"Total {job_name}"
    ws[f'B{footer_row}'].font = Font(bold=True)
    total_col_letter = chr(ord('A') + len(generated_months) + 4)
    ws[f'{total_col_letter}{footer_row}'].number_format = '#,##0.00'

    footer_rows.append(footer_row)

    for i, month in enumerate(generated_months, start=0):
        month_col_letter = chr(ord(first_month_col) + i)
        month_cell_range = f'{month_col_letter}{group_start_row}:{month_col_letter}{end_row}'
        month_cell = f'{month_col_letter}{footer_row}'
        ws[month_cell] = f'=SUM({month_cell_range})'
        ws[month_cell].number_format = '#,##0.00_);(#,##0.00)'

        month_datetime = datetime.strptime(month, '%b-%y')
        if month_datetime <= last_month_dt:
            ws[month_cell].fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
        else:
            ws[month_cell].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws[month_cell].font = Font(bold=True)

    total_col_letter = chr(ord(first_month_col) + len(generated_months))
    ws[f'{total_col_letter}{footer_row}'] = f'=SUM({first_month_col}{footer_row}:{month_col_letter}{footer_row})'
    ws[f'{total_col_letter}{footer_row}'].font = Font(bold=True)

    return footer_row

def find_employee_row(ws, employee_name, labor_category_name):
    for row in range(1, ws.max_row + 1):
        if ws[f'B{row}'].value == employee_name and ws[f'A{row}'].value == labor_category_name:
            return row
    return None

def add_total_for_row(ws, row, start_col, end_col):
    total_formula = f"=SUM({start_col}{row}:{end_col}{row})"
    set_cell(ws, f'{chr(ord(end_col) + 1)}{row}', total_formula, bold=True)
    ws[f'{chr(ord(end_col) + 1)}{row}'].number_format = '#,##0.00'

#categorize months
def categorize_months(dc_start_year, dc_end_year, last_month):
    all_actual_months = []
    all_forecast_months = []

    for year in range(dc_start_year, dc_end_year + 1): 
        actual_months = []
        forecast_months = []

        for month in range(1, 13):
            month_date = datetime(year, month, 1)
            month_str = month_date.strftime("%b-%y")

            if month_date <= last_month:
                actual_months.append(month_str)
            else:
                forecast_months.append(month_str)
                
        all_actual_months.extend(actual_months)
        all_forecast_months.extend(forecast_months)
    return all_actual_months, all_forecast_months
 