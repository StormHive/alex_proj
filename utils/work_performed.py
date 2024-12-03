from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from utils.helper_functions import categorize_months, list_contract_months, convert_to_datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
      
def create_work_performed_spreadsheet(contract_id, data1, data2, data3, lookup_month_info, filename, pop_start_date, pop_end_date, work_year, dc_start_year , dc_end_year , last_month_str):
    try:
        wb = Workbook()
        wb.remove(wb.active)
        last_month = convert_to_datetime(last_month_str)
        
        for year in range(dc_start_year, dc_end_year + 1):
            ws2 = wb.create_sheet(title=f"{year} Work Performed")
        
            
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") 
            header_font = Font(bold=True, color="FFFFFF") 
  
            for r_idx, row in enumerate(dataframe_to_rows(data2, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:  
                        cell.fill = header_fill
                        cell.font = header_font
            
            period_numbers = data3['PeriodNumber'].unique()
            for period_number in period_numbers:
                hrs_data = data3[data3['PeriodNumber'] == period_number]
                if hrs_data.empty:
                    print(f"No data found for period number: {period_number}")
                    continue
                
                period_start_date = hrs_data['StartDate'].min().strftime("%Y-%m-%d")
                period_end_date = hrs_data['EndDate'].max().strftime("%Y-%m-%d")

                generated_months = list_contract_months(period_start_date, period_end_date)

        wb.save(filename)
        print(f"Data successfully saved to {filename}")

    except Exception as e:
        print(f"Error saving data to Excel: {e}")


