from openpyxl import Workbook
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from utils.helper_functions import (set_cell, find_employee_row, list_contract_months, add_total_for_row, 
create_footer_row,set_font_for_sheet,add_footer_rows)

#create budget sheet
def create_budget_spreadsheet(wb, headers_data, data, lookup_month_info, last_month_dt, PeriodNumber, hrs_sheet_title):
    global footer_rows
    footer_rows = []

    sheet_title = f"OY{PeriodNumber} Budget" if PeriodNumber >= 1 else "Base Budget"
    ws = wb.create_sheet(title=sheet_title)
    set_font_for_sheet(ws)

    bold_font = Font(bold=True)
    salmon_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    headers = [
        ("B1", "Prime Contract #", headers_data['prime contract #'].iloc[0]),
        ("B2", "Task #", headers_data['task #'].iloc[0]),
        ("B3", "Contract Name", headers_data['contract name'].iloc[0]),
        ("B4", "POP", headers_data['pop'].iloc[0]),
        ("B5", "DM", headers_data['dm'].iloc[0]),
        ("B6", "PM", headers_data['pm'].iloc[0])
    ]
    
    for cell, header, value in headers:
        set_cell(ws, cell, header, bold=True)
        set_cell(ws, cell.replace('B', 'C'), value, bold=True)

    set_cell(ws, 'B8', "BY T&M", bold=True)
    set_cell(ws, 'C8', "To be determined", bold=True)
    set_cell(ws, 'A12', "LCAT CD", bold=True)
    set_cell(ws, 'B12', "Labor Category", bold=True)
    set_cell(ws, 'C12', "Name", bold=True)
    set_cell(ws, 'D12', "Rate", bold=True)

    pop_end_date = pd.Timestamp(headers_data['pop'].iloc[0].split(' - ')[1])   
    pop_start_date = pd.Timestamp(headers_data['pop'].iloc[0].split(' - ')[0])
    months = list_contract_months(pop_start_date.strftime("%Y-%m-%d"), pop_end_date.strftime("%Y-%m-%d"))
    generated_months = months

    total_col_index = len(generated_months) + 5 
    for col in range(2, total_col_index + 1):  
        ws.cell(row=11, column=col).fill = salmon_fill
        ws.cell(row=12, column=col).fill = salmon_fill

    for i, month in enumerate(generated_months, start=5):
        month_col_letter = chr(ord('A') + i - 1)
        set_cell(ws, f'{month_col_letter}11', month)
        ws[f'{month_col_letter}11'].number_format = 'mmm-yy'
        month_datetime = datetime.strptime(month, '%b-%y')
        work_days_available = lookup_month_info[
            (lookup_month_info['Month'] == month_datetime.month) &
            (lookup_month_info['Year'] == month_datetime.year)
        ]['WorkDaysAvailable'].values

    total_col_index = len(generated_months) + 5
    total_col_letter = chr(ord('A') + total_col_index - 1)
    set_cell(ws, f'{total_col_letter}11', "Total", bold=True)

    start_row = 13
    start_row = create_budget_employee_rows(ws, data, lookup_month_info, start_row, generated_months, last_month_dt, hrs_sheet_title, PeriodNumber)
    add_additional_rows(ws, start_row, generated_months, last_month_dt)
    return wb

#create foother
def create_footer_row(ws, group_start_row, end_row, generated_months, last_month_dt, job_name):
    footer_row = end_row + 1
    first_month_col = 'E'

    group_job_name = ws[f'B{group_start_row}'].value.split(":")[0]
    ws[f'B{footer_row}'].value = f"Total {job_name}"
    ws[f'B{footer_row}'].font = Font(bold=True)
    total_col_letter = chr(ord('A') + len(generated_months) + 4)
    ws[f'{total_col_letter}{footer_row}'].number_format = '#,##0.00'
    footer_rows.append(footer_row)

    for i, month in enumerate(generated_months, start=0):
        month_col_letter = chr(ord(first_month_col) + i)
        month_cell_range = f'{month_col_letter}{group_start_row}:{month_col_letter}{end_row}'
        month_cell = f'{month_col_letter}{footer_row}'
        ws[month_cell] = f'=SUM({month_cell_range})'
        ws[month_cell].number_format = '#,##0.00_);(#,##0.00)'

        month_datetime = datetime.strptime(month, '%b-%y')
        if month_datetime <= last_month_dt:
            ws[month_cell].fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
        else:
            ws[month_cell].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws[month_cell].font = Font(bold=True)

    total_col_letter = chr(ord(first_month_col) + len(generated_months))
    ws[f'{total_col_letter}{footer_row}'] = f'=SUM({first_month_col}{footer_row}:{month_col_letter}{footer_row})'
    ws[f'{total_col_letter}{footer_row}'].font = Font(bold=True)
    return footer_row

def create_budget_employee_rows(ws, data, lookup_month_info, start_row, generated_months, last_month_dt, hrs_sheet_title, PeriodNumber):
    current_job = None
    current_employee = None
    group_start_row = start_row
    hrs_ws = ws.parent[hrs_sheet_title]

    grouped_data = data.groupby('JobTitle')
    for job_name, group in grouped_data:
        set_cell(ws, f'B{start_row}', job_name.split(":")[0], bold=True)
        for col in range(1, len(generated_months) + 6):
            ws.cell(row=start_row, column=col).fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

        start_row += 1  
        current_employee = None
        for _, row in group.iterrows():
            if row['EmployeeName'] != current_employee:
                current_employee = row['EmployeeName']
                set_cell(ws, f'A{start_row}', row['LaborCategoryCode'])
                set_cell(ws, f'B{start_row}', row['LaborCategoryName'])
                if row['EmployeeName'] == 'TBD':
                    set_cell(ws, f'C{start_row}', f"TBD - {row['Company']} - {row['PreviousEmployeeLastName']} Replacement")
                else:
                    set_cell(ws, f'C{start_row}', row['EmployeeName'])
                    set_cell(ws, f'D{start_row}', row['LaborCategoryRate'])
                    ws[f'D{start_row}'].number_format = '#,##0.00'

                employee_row = find_employee_row(hrs_ws, row['EmployeeName'], row['LaborCategoryName'])
                if employee_row is None:
                    continue

                for month_col, month in enumerate(generated_months, start=5):
                    month_col_letter = chr(ord('A') + month_col - 1)
                    month_datetime = datetime.strptime(month, '%b-%y')
                    cell_value = f"='{hrs_sheet_title}'!{month_col_letter}{employee_row}*$D{start_row}"
                    set_cell(ws, f'{month_col_letter}{start_row}', cell_value, fill=PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))
                    ws[f'{month_col_letter}{start_row}'].number_format = '#,##0.00'

                # total for the employee row
                total_col_letter = chr(ord('A') + len(generated_months) + 4)
                total_formula = f"=SUM(E{start_row}:{month_col_letter}{start_row})"
                set_cell(ws, f'{total_col_letter}{start_row}', total_formula, bold=True)
                ws[f'{total_col_letter}{start_row}'].number_format = '#,##0.00'
                start_row += 1  

        set_cell(ws, f'B{start_row}', f'Total {job_name.split(":")[0]}', bold=True)
        for i, month in enumerate(generated_months, start=5):
            month_col_letter = chr(ord('A') + i - 1)
            total_group_formula = f"=SUM({month_col_letter}{group_start_row}:{month_col_letter}{start_row - 1})"
            set_cell(ws, f'{month_col_letter}{start_row}', total_group_formula, bold=True)
            ws[f'{month_col_letter}{start_row}'].number_format = '#,##0.00'

        # total for group
        total_col_letter = chr(ord('A') + len(generated_months) + 4)
        total_group_formula = f"=SUM(E{start_row}:{month_col_letter}{start_row})"
        set_cell(ws, f'{total_col_letter}{start_row}', total_group_formula, bold=True)
        ws[f'{total_col_letter}{start_row}'].number_format = '#,##0.00'
        start_row += 1
        group_start_row = start_row

    return start_row

def add_additional_rows(ws, start_row, generated_months, last_month_dt):
    current_row = start_row + 2
    start_col_letter = 'E'
    end_col_letter = chr(ord(start_col_letter) + len(generated_months) - 1)
    total_col_letter = chr(ord(end_col_letter) + 1)

    # Add Total Labor Revenue
    set_cell(ws, f'B{current_row}', "Total Labor Revenue", bold=True)
    for i, month in enumerate(generated_months, start=5):
        month_col_letter = chr(ord('A') + i - 1)
        month_cell = f'{month_col_letter}{current_row}'
        total_labor_revenue_total_formula = f"=SUM({month_col_letter}{start_row - len(generated_months)}:{month_col_letter}{start_row})"
        set_cell(ws, month_cell, total_labor_revenue_total_formula, bold=True)
        ws[month_cell].number_format = '#,##0.00'
        ws[month_cell].fill = PatternFill(
            start_color="92CDDC" if datetime.strptime(generated_months[i-5], '%b-%y') <= last_month_dt else "FFFFFF",
            end_color="FFFFFF", fill_type="solid"
        )
    
    def add_total_for_row(ws, row, start_col, end_col):
        total_formula = f"=SUM({start_col}{row}:{end_col}{row})"
        set_cell(ws, f'{chr(ord(end_col) + 1)}{row}', total_formula, bold=True)
        ws[f'{chr(ord(end_col) + 1)}{row}'].number_format = '#,##0.00'

    add_total_for_row(ws, current_row, start_col_letter, end_col_letter)
    total_labor_revenue_row = current_row
    current_row += 1

    set_cell(ws, f'B{current_row}', "ODCs", bold=True, fill=PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"))
    set_cell(ws, f'C{current_row}', "M&S Rate", bold=True, fill=PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"))
    set_cell(ws, f'D{current_row}', 0, bold=True, fill=PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"))
    ws[f'D{current_row}'].number_format = '0.00%'
    for i, month in enumerate(generated_months, start=5):
        month_col_letter = chr(ord('A') + i - 1)
        set_cell(ws, f'{month_col_letter}{current_row}', None, fill=PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"))
    add_total_for_row(ws, current_row, start_col_letter, end_col_letter)
    current_row += 1

    set_cell(ws, f'B{current_row}', "Software", bold=True)
    add_total_for_row(ws, current_row, start_col_letter, end_col_letter)
    current_row += 1

    set_cell(ws, f'B{current_row}', "Archer", bold=True)
    add_total_for_row(ws, current_row, start_col_letter, end_col_letter)
    current_row += 3

    set_cell(ws, f'B{current_row}', "Total ODC Revenue", bold=True)
    for i, month in enumerate(generated_months, start=5):
        month_col_letter = chr(ord('A') + i - 1)
        month_cell = f'{month_col_letter}{current_row}'
        set_cell(ws, month_cell, 0, bold=True)
        ws[month_cell].number_format = '#,##0.00'
        ws[month_cell].fill = PatternFill(start_color="92CDDC" if datetime.strptime(generated_months[i-5], '%b-%y') <= last_month_dt else "FFFFFF", end_color="FFFFFF", fill_type="solid")
    add_total_for_row(ws, current_row, start_col_letter, end_col_letter)
    total_odc_revenue_row = current_row
    current_row += 1

    set_cell(ws, f'B{current_row}', "Travel", bold=True, fill=PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"))
    set_cell(ws, f'C{current_row}', None, fill=PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"))
    set_cell(ws, f'D{current_row}', None, fill=PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"))
    for i, month in enumerate(generated_months, start=5):
        month_col_letter = chr(ord('A') + i - 1)
        set_cell(ws, f'{month_col_letter}{current_row}', None, fill=PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"))
    add_total_for_row(ws, current_row, start_col_letter, end_col_letter)
    current_row += 1

    set_cell(ws, f'B{current_row}', "Total Travel Revenue", bold=True)
    for i, month in enumerate(generated_months, start=5):
        month_col_letter = chr(ord('A') + i - 1)
        month_cell = f'{month_col_letter}{current_row}'
        set_cell(ws, month_cell, 0, bold=True)
        ws[month_cell].number_format = '#,##0.00'
        ws[month_cell].fill = PatternFill(start_color="92CDDC" if datetime.strptime(generated_months[i-5], '%b-%y') <= last_month_dt else "FFFFFF", end_color="FFFFFF", fill_type="solid")
    add_total_for_row(ws, current_row, start_col_letter, end_col_letter)
    total_travel_revenue_row = current_row
    current_row += 1

    set_cell(ws, f'B{current_row}', "Total Revenue", bold=True)
    for i, month in enumerate(generated_months, start=5):
        month_col_letter = chr(ord('A') + i - 1)
        month_cell = f'{month_col_letter}{current_row}'
        total_formula = f"={month_col_letter}{total_labor_revenue_row} + {month_col_letter}{total_odc_revenue_row} + {month_col_letter}{total_travel_revenue_row}"
        set_cell(ws, month_cell, total_formula, bold=True)
        ws[month_cell].number_format = '#,##0.00'
        ws[month_cell].fill = PatternFill(
            start_color="92CDDC" if datetime.strptime(generated_months[i-5], '%b-%y') <= last_month_dt else "FFFFFF",
            end_color="FFFFFF", fill_type="solid"
        )
    add_total_for_row(ws, current_row, start_col_letter, end_col_letter)
    current_row += 1