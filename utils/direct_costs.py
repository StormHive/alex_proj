
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import openpyxl.utils.cell as cell_utils
import utils.helper_functions as helper_functions
import pandas as pd
from openpyxl.utils import get_column_letter
from utils.helper_functions import (get_direct_costs_data, get_hrs_and_budget_data, convert_to_datetime, 
check_month_period_number, sort_key, add_columns,get_period_number_data)

def create_direct_costs_spreadsheet(data, ws, wb, work_year, actual_months, forecast_months, year, last_month_str, contract_id, dc_start_year, dc_end_year, sql_result):
    formula_array = [] 
    try:
        last_month_dt = datetime.strptime(last_month_str, "%m/%Y")
        pop_ids = helper_functions.get_unique_pop_ids()

        # Direct Costs Data
        data1 = get_direct_costs_data(helper_functions.db_uri, contract_id, work_year)
        all_data3 = []  
        data2 = helper_functions.get_work_performed_data(helper_functions.db_uri, contract_id, work_year)
        
        for pop_id in pop_ids['pop_id']:
            data3, lookup_month_info, pop_start_date, pop_end_date = get_hrs_and_budget_data(helper_functions.db_uri, pop_id, last_month_str)
            if data3 is not None and not data3.empty and not data3.isna().all().all():  
                all_data3.append(data3)
            else:
                print(f"No data found for pop_id: {pop_id}")

        all_data3_filtered = [df.dropna(axis=1, how='all') for df in all_data3 if not df.empty and not df.isna().all().all()]
        if all_data3_filtered:
            combined_data3 = pd.concat(all_data3_filtered, ignore_index=True)
        else:
            print("No valid data found")
            combined_data3 = pd.DataFrame() 

        period_data_list = get_period_number_data(combined_data3)

        # sheet title
        ws.title = f"{year} Direct Costs"
        direct_cost_sheet_name = ws.title
        work_performed_sheet_name = f"{year} Work Performed"

        # Header row and columns
        ws['B9'] = "Hours"
        column_order = [
            'Employee ID', 'EmployeeName', 'Billable Rate', 'Direct Rate', 
            'CM%', 'OI%',  '% Direct', 'Employee / Sub'
        ]

        # Filter employees
        regular_employees = data[(data['IsTbd'] == False) | (data['company_id'] != 2)]
        tbd_employees = data[(data['IsTbd'] == True) & (data['company_id'] == 2)]
       
        # Order columns
        regular_employees = regular_employees[column_order]
        tbd_employees = tbd_employees[column_order]

        # Populate the headers
        headers = list(regular_employees.columns) 
        for col_num, header in enumerate([''] + headers, 1):
            ws.cell(row=8, column=col_num, value=header)

        # Header
        header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num in range(2, len(headers) + 2):
            cell = ws.cell(row=8, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        # Column indices
        col_indices = {header: idx + 2 for idx, header in enumerate(headers)}
        negative_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        last_row_idx = 0
       
        # populate_rows_for_group 
        def populate_rows_for_group(employee_group, start_row, is_tbd=False):
            nonlocal last_row_idx              
            for r_idx, row in enumerate(dataframe_to_rows(employee_group, index=False, header=False), start=start_row):
                last_row_idx = r_idx  
                for col_num, value in enumerate([''] + row, 1):
                    ws.cell(row=r_idx, column=col_num, value=value)

                # Set Employee / Sub column value
                employee_sub_col_idx = col_indices['Employee / Sub']
                if is_tbd:
                    ws.cell(row=r_idx, column=employee_sub_col_idx, value="Subcontractor Labor")
                else:
                    ws.cell(row=r_idx, column=employee_sub_col_idx, value="Employee Labor")

                # Format columns
                for col_name, col_index in col_indices.items():
                    cell = ws.cell(row=r_idx, column=col_index)
                    if col_name in ['Billable rate', 'DirectRate', 'CM%', 'OI%', '% Direct']:
                        cell.fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
                        if col_name in ['Billable rate', 'DirectRate']:
                            cell.number_format = '$#,##0.00'
                        elif col_name in ['CM%', 'OI%', '% Direct']:
                            cell.number_format = '0.00%'
                        if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value < 0:
                            cell.fill = negative_fill

                # Get percentages and formulas
                percentages_CM_formula = helper_functions.get_contract_percentages(contract_id)
                fringe_percentage = percentages_CM_formula.iloc[0]['FringePercentage'] * 100
                oh_percentage = percentages_CM_formula.iloc[0]['OHPercentage'] * 100
                ms_percentage = percentages_CM_formula.iloc[0]['MSPercentage'] * 100
                ga_percentage = percentages_CM_formula.iloc[0]['GAPercentage'] * 100

                EE_Wrap_OI_value = round((1 * (1 + fringe_percentage / 100) * (1 + oh_percentage / 100) * (1 + ga_percentage / 100)), 2)

                # Apply formula
                CM_formula = f"=IFERROR(({cell_utils.get_column_letter(col_indices['Billable Rate'])}{r_idx} - ({cell_utils.get_column_letter(col_indices['Direct Rate'])}{r_idx} * (1 + {fringe_percentage / 100}))) / {cell_utils.get_column_letter(col_indices['Billable Rate'])}{r_idx}, 0)"
                ws[f'{cell_utils.get_column_letter(col_indices['CM%'])}{r_idx}'].value = CM_formula

                OI_formula = f"=IFERROR(({cell_utils.get_column_letter(col_indices['Billable Rate'])}{r_idx} - ({cell_utils.get_column_letter(col_indices['Direct Rate'])}{r_idx} * {EE_Wrap_OI_value})) / {cell_utils.get_column_letter(col_indices['Billable Rate'])}{r_idx}, 0)"
                ws[f'{cell_utils.get_column_letter(col_indices['OI%'])}{r_idx}'].value = OI_formula

                countif_formula = f"=COUNTIF(C:C, C{r_idx})"
                ws[f'A{r_idx}'].value = countif_formula
                
                # Store formulas in array
                formula_array.append({
                    'row': r_idx,
                    'CM_formula': CM_formula,
                    'OI_formula': OI_formula,
                    'countif_formula': countif_formula
                })
            return last_row_idx  
        
        # Populate regular employees
        last_row_regular = populate_rows_for_group(regular_employees, start_row=10, is_tbd=False)

        # Empty rows between regular and TBD employees
        empty_row_count = 6
        empty_row_start = last_row_regular + 1
        for _ in range(empty_row_count):
            ws.append([''] * (len(headers) + 1))

        # Populate TBD employees
        last_row_tbd = populate_rows_for_group(tbd_employees, start_row=empty_row_start + empty_row_count, is_tbd=True)

        # Combine actual and forecast months
        combined_months = actual_months + forecast_months
        combined_months.sort(key=sort_key)

        # Map Actual/Forecast months 
        month_labels = {}
        for month in combined_months:
            is_actual = month in actual_months
            category = "Actual" if is_actual else "Forecast"
            month_labels[month] = category

        # First section for months
        header_start_col = len(headers) + 2
        for col_num, month in enumerate(combined_months, start=header_start_col):
            is_actual = month in actual_months
            category = "Actual" if is_actual else "Forecast"
            ws.cell(row=6, column=col_num, value=category)
            ws.cell(row=7, column=col_num, value=month)
            ws.cell(row=6, column=col_num).font = Font(bold=True)
            ws.cell(row=7, column=col_num).font = header_font 
            ws.cell(row=7, column=col_num).fill = header_fill
            ws.cell(row=4, column=col_num, value=month)

            month_date = convert_to_datetime(month)
            forecast_fill_color = "ADD8E6"  
            actual_fill_color = "B0C4DE"

            # get all sheet
            all_sheet_names = wb.sheetnames

            direct_cost_sheets = [sheet for sheet in all_sheet_names if 'Direct Cost' in sheet]
            hrs_sheets = [sheet for sheet in all_sheet_names if 'HRS' in sheet and ('oy' in sheet.lower() or 'base' in sheet.lower())]

            work_performed_df = sql_result[['EmployeeName', 'month_match', 'Work year']].drop_duplicates()
            for r_idx in range(10, last_row_idx + 1):
                if empty_row_start <= r_idx < empty_row_start + empty_row_count:
                    continue

                employee_name = ws.cell(row=r_idx, column=col_indices['EmployeeName']).value
                for col_num in range(10, ws.max_column + 1):
                    month = ws.cell(row=7, column=col_num).value
                    if not month:
                        continue
                    
                    month_abr = month[:3]
                    month_date = convert_to_datetime(month)

                    if month_date <= last_month_dt:
                        if not work_performed_df[
                            (work_performed_df['EmployeeName'] == employee_name) &
                            (work_performed_df['month_match'] == f"{month_abr}-{str(year)[-2:]}") &
                            (work_performed_df['Work year'] == year)
                        ].empty:
                            formula = (
                                f"=SUMIFS('{work_performed_sheet_name}'!$P:$P, "
                                f"'{work_performed_sheet_name}'!$A:$A, \"*{month[:3]}*\", "
                                f"'{work_performed_sheet_name}'!$L:$L, '{direct_cost_sheet_name}'!$B{r_idx})"
                            )
                            ws.cell(row=r_idx, column=col_num, value=formula)
                            ws.cell(row=r_idx, column=col_num).fill = PatternFill(start_color=actual_fill_color, end_color=actual_fill_color, fill_type="solid")
                        else:
                            ws.cell(row=r_idx, column=col_num, value=0)
                            ws.cell(row=r_idx, column=col_num).number_format = '0'
                            ws.cell(row=r_idx, column=col_num).fill = PatternFill(start_color=actual_fill_color, end_color=actual_fill_color, fill_type="solid")
                    else:
                        dynamic_period_number = helper_functions.check_month_period_number(month, period_data_list, default_period_number=-1)

                        if dynamic_period_number == -1:
                            ws.cell(row=r_idx, column=col_num, value=0)
                        elif dynamic_period_number == 0:
                            hrs_sheet_name = "Base HRS"
                        elif dynamic_period_number >= 1:
                            hrs_sheet_name = f"OY{dynamic_period_number} HRS"

                        for col in range(10, ws.max_column + 1):  
                            forecast_header = ws.cell(row=6, column=col).value

                            if "Forecast" in forecast_header: 
                                month_value = ws.cell(row=7, column=col).value  
                                
                                # iterate HRS sheets
                                match_found = False
                                for hrs_sheet_name in hrs_sheets:
                                    hrs_ws = wb[hrs_sheet_name]
                                    for hrs_row in range(10, hrs_ws.max_row + 1):
                                        hrs_employee_name = hrs_ws.cell(row=hrs_row, column=2).value

                                        if hrs_employee_name and hrs_employee_name.strip().lower() == employee_name.strip().lower():
                                            for hrs_col in range(5, hrs_ws.max_column + 1):
                                                hrs_month_value = hrs_ws.cell(row=9, column=hrs_col).value

                                                if hrs_month_value == month_value:
                                                    col_letter = get_column_letter(hrs_col)
                                                    
                                                    forecast_formula = (
                                                        f"=ROUND(SUMIFS('{hrs_sheet_name}'!{col_letter}:{col_letter}, "
                                                        f"'{hrs_sheet_name}'!$B:$B, '{direct_cost_sheet_name}'!$C{r_idx}), 1)"
                                                    )
                                                    
                                                    ws.cell(row=r_idx, column=col, value=forecast_formula)
                                                    ws.cell(row=r_idx, column=col).fill = PatternFill(start_color=forecast_fill_color, end_color=forecast_fill_color, fill_type="solid")

                                                    match_found = True
                                                    break
                                    if match_found:
                                        break

                                if not match_found:
                                    ws.cell(row=r_idx, column=col, value=0)
                                    ws.cell(row=r_idx, column=col).number_format = '0'
                                    ws.cell(row=r_idx, column=col).fill = PatternFill(start_color=forecast_fill_color, end_color=forecast_fill_color, fill_type="solid")

        # Hours & Total columns 
        total_col_idx = header_start_col + len(combined_months) + 1
        ws.cell(row=6, column=total_col_idx, value="Hours").font = Font(bold=True)
        ws.cell(row=7, column=total_col_idx, value="Total").font = Font(bold=True)
        ws.cell(row=7, column=total_col_idx).fill = header_fill
        ws.cell(row=7, column=total_col_idx).font = header_font
        total_column = total_col_idx  
        total_row = 7  
        
        # Hours and Total
        for r_idx in range(10, last_row_idx + 1):
            if empty_row_start <= r_idx < empty_row_start + empty_row_count:
                continue

            actual_month_start_col = header_start_col
            forecast_month_end_col = total_col_idx - 1

            hours_formula = f"=SUM({cell_utils.get_column_letter(actual_month_start_col)}{r_idx}:{cell_utils.get_column_letter(forecast_month_end_col)}{r_idx})"
            ws.cell(row=r_idx, column=total_col_idx, value=hours_formula)

         # Second section
        second_section_start_col = ws.max_column + 2 
        for col_num, month in enumerate(combined_months, start=second_section_start_col):
            category = "Actual" if month in actual_months else "Forecast"
            ws.cell(row=6, column=col_num, value=category)  
            ws.cell(row=7, column=col_num, value=month) 
            ws.cell(row=6, column=col_num).font = Font(bold=True)
            ws.cell(row=7, column=col_num).font = header_font
            ws.cell(row=7, column=col_num).fill = header_fill

        # Total/Labor
        sectiontwo_total_col_idx = second_section_start_col + len(combined_months)
        ws.cell(row=6, column=sectiontwo_total_col_idx, value="Labor").font = Font(bold=True)
        ws.cell(row=7, column=sectiontwo_total_col_idx, value="Total").font = Font(bold=True)
        ws.cell(row=7, column=sectiontwo_total_col_idx).fill = header_fill
        ws.cell(row=7, column=sectiontwo_total_col_idx).font = header_font

        for r_idx in range(10, last_row_idx + 1):
            if empty_row_start <= r_idx < empty_row_start + empty_row_count:
                continue

            labor_formula = f"=SUM({cell_utils.get_column_letter(second_section_start_col)}{r_idx}:{cell_utils.get_column_letter(sectiontwo_total_col_idx - 1)}{r_idx})"
            ws.cell(row=r_idx, column=sectiontwo_total_col_idx, value=labor_formula)
            ws.cell(row=r_idx, column=sectiontwo_total_col_idx).number_format = '$ #,##0.00'

        #[Year] Raise
        year_raise_col_idx = sectiontwo_total_col_idx + 2 
        ws.cell(row=9, column=year_raise_col_idx, value="[Year] Raise").font = Font(bold=True)
        year_raise_values = [30.76, 44.72, 56.09, 22.58, 30.47, 50.87, 25.01]
        for row_num, value in enumerate(year_raise_values, start=10):
            ws.cell(row=row_num, column=year_raise_col_idx, value=value).number_format = '$ #,##0.00' 

        for col in range(1, ws.max_column + 1):
            if ws.cell(row=9, column=col).value == "[Year] Raise":
                reference_col = cell_utils.get_column_letter(col)
                break
        new_col = cell_utils.get_column_letter(ws.max_column + 1)

        for row in range(10, last_row_idx + 1):
            if empty_row_start <= row < empty_row_start + empty_row_count:
                continue  
            
            ws[f'{new_col}{row}'] = f"=IFERROR({reference_col}{row}/E{row}-1,0)"
            ws[f'{new_col}{row}'].number_format = '0.0%'
        ws[f'{new_col}9'] = " "

        month_date = convert_to_datetime(month)

        # Loop second section
        for r_idx in range(10, last_row_idx + 1):
            if empty_row_start <= r_idx < empty_row_start + empty_row_count:
                continue  
            
            for col_idx, month in enumerate(combined_months, start=second_section_start_col):
                category = month_labels[month]

                if category == "Actual":
                    section_2_formula = (
                        f"=SUMIFS('{work_performed_sheet_name}'!$Q:$Q,'{work_performed_sheet_name}'!$A:$A, \"*{month[:3]}*\", "
                        f"'{work_performed_sheet_name}'!$L:$L, '{direct_cost_sheet_name}'!$B{r_idx})" )
                    
                    ws.cell(row=r_idx, column=col_idx, value=section_2_formula)
                    ws.cell(row=r_idx, column=col_idx).fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
                    ws.cell(row=r_idx, column=col_idx).number_format = '$* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

                elif category == "Forecast":
                    if month_date <= last_month_dt:
                        section_2_formula = "0"
                    else:
                        match_forecast_by_month = cell_utils.get_column_letter(
                            header_start_col + len(actual_months) + (col_idx - (second_section_start_col + len(actual_months)))
                        )
                        year_raise_col_letter = cell_utils.get_column_letter(year_raise_col_idx)
                        section_2_formula = f"={match_forecast_by_month}{r_idx}*${year_raise_col_letter}${r_idx}"
                    
                    ws.cell(row=r_idx, column=col_idx, value=section_2_formula)
                    ws.cell(row=r_idx, column=col_idx).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    ws.cell(row=r_idx, column=col_idx).number_format = '$* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

        # End column for regular employees
        end_col_regular = header_start_col + len(actual_months) + len(forecast_months) - 1

        regular_actual_forecast_footer_row = last_row_regular + 2 
        tbd_actual_forecast_footer_row = last_row_tbd + 2

        col_num = header_start_col
        while col_num <= ws.max_column:
            col_letter = cell_utils.get_column_letter(col_num)
            is_empty = all(ws.cell(row=r, column=col_num).value is None for r in range(7, 9))  
            if not is_empty:

                # Footer for regular employees
                sum_footer_regular = f"=SUM({col_letter}10:{col_letter}{last_row_regular})"
                ws.cell(row=regular_actual_forecast_footer_row, column=col_num, value=sum_footer_regular)
                ws.cell(row=regular_actual_forecast_footer_row, column=col_num).font = Font(bold=True)

                # Footer for TBD employees
                sum_footer_tbd = f"=SUM({col_letter}{empty_row_start + empty_row_count + 1}:{col_letter}{last_row_tbd})"
                ws.cell(row=tbd_actual_forecast_footer_row, column=col_num, value=sum_footer_tbd)
                ws.cell(row=tbd_actual_forecast_footer_row, column=col_num).font = Font(bold=True)

                if col_num >= second_section_start_col:  
                    ws.cell(row=regular_actual_forecast_footer_row, column=col_num).number_format = '$ #,##0.00'  
                    ws.cell(row=tbd_actual_forecast_footer_row, column=col_num).number_format = '$ #,##0.00'      
            col_num += 1

        # empty rows
        def add_empty_rows(start_row, num_rows, num_columns):
            for i in range(num_rows):
                for col_num in range(1, num_columns + 1):
                    ws.cell(row=start_row + i, column=col_num).value = None

        # Total Employee Labor
        EmployeeLabor_total_label_row = last_row_regular + 2
        Total_EmployeeLabor = ws.cell(row=EmployeeLabor_total_label_row, column=col_indices['Employee / Sub'], value="Total Employee Labor")
        Total_EmployeeLabor.font = Font(bold=True)

        # Total Employee Labor
        num_empty_rows = 2
        add_empty_rows(EmployeeLabor_total_label_row + 1, num_empty_rows, len(headers) + 1)

        # Total Subcontractor Labor
        TotalSubcontractor_total_label_row = last_row_tbd + 2
        Total_Subcontractor = ws.cell(row=TotalSubcontractor_total_label_row, column=col_indices['Employee / Sub'], value="Total Subcontractor Labor")
        Total_Subcontractor.font = Font(bold=True)

        # empty rows after Total Subcontractor Labor
        add_empty_rows(TotalSubcontractor_total_label_row + 1, num_empty_rows, len(headers) + 1)

        # Direct Software, Billable ODC, Non-Billable ODCs, Total ODCs
        Billable_total_label_row = TotalSubcontractor_total_label_row + num_empty_rows + 1
        ws.cell(row=Billable_total_label_row, column=col_indices['Employee / Sub'], value="Direct Software")
        ws.cell(row=Billable_total_label_row + 1, column=col_indices['Employee / Sub'], value="Billable ODC")
        ws.cell(row=Billable_total_label_row + 2, column=col_indices['Employee / Sub'], value="Non-Billable ODCs")
        total_odcs = ws.cell(row=Billable_total_label_row + 3, column=col_indices['Employee / Sub'], value="Total ODCs")
        total_odcs.font = Font(bold=True)

        # empty rows after Total ODCs
        add_empty_rows(Billable_total_label_row + 4, num_empty_rows, len(headers) + 1)

        # Total Direct Costs
        Direct_total_label_row = Billable_total_label_row + num_empty_rows + 4
        Total_Direct_Costs = ws.cell(row=Direct_total_label_row, column=col_indices['Employee / Sub'], value="Total Direct Costs")
        Total_Direct_Costs.font = Font(bold=True)

        # empty rows after Total Direct Costs
        add_empty_rows(Direct_total_label_row + 1, num_empty_rows, len(headers) + 1)
        Employee_total_label_row = Direct_total_label_row + num_empty_rows + 1
        
        # Employee Labor
        ws.cell(row=Employee_total_label_row, column=col_indices['Employee / Sub'], value="Employee Labor")
        col_num = header_start_col

        while col_num <= ws.max_column:
            col_letter = cell_utils.get_column_letter(col_num)
            is_empty = all(ws.cell(row=r, column=col_num).value is None for r in range(7, 9))
            if not is_empty:
                total_employee_labor_row = regular_actual_forecast_footer_row

                reference_formula = f"={col_letter}{total_employee_labor_row}"
                ws.cell(row=Employee_total_label_row, column=col_num).value = reference_formula
                ws.cell(row=Employee_total_label_row, column=col_num).font = Font(bold=False)
                if col_num >= second_section_start_col:
                    ws.cell(row=Employee_total_label_row, column=col_num).number_format = '$* #,##0_);($* (#,##0);$* 0_);_(@_)'
            
            col_num += 1  

        # Subcontractor Labor
        subcontractor_total_label_row = Employee_total_label_row + 1
        ws.cell(row=subcontractor_total_label_row, column=col_indices['Employee / Sub'], value="Subcontractor Labor")
        col_num = header_start_col

        while col_num <= ws.max_column:
            col_letter = cell_utils.get_column_letter(col_num)
            is_empty = all(ws.cell(row=r, column=col_num).value is None for r in range(7, 9))
            if not is_empty:
                reference_formula = f"={col_letter}{tbd_actual_forecast_footer_row}" 
                ws.cell(row=subcontractor_total_label_row, column=col_num).value = reference_formula
                ws.cell(row=subcontractor_total_label_row, column=col_num).font = Font(bold=False)
                if col_num >= second_section_start_col:
                    ws.cell(row=subcontractor_total_label_row, column=col_num).number_format = '$* #,##0_);($* (#,##0);$* 0_);_(@_)'
            
            col_num += 1 

        #Direct Software
        ws.cell(row=Employee_total_label_row + 2, column=col_indices['Employee / Sub'], value="Direct Software")
        ws.cell(row=Employee_total_label_row + 3, column=col_indices['Employee / Sub'], value="Billable ODC")

        # Non-Billable ODCs
        ws.cell(row=Employee_total_label_row + 4, column=col_indices['Employee / Sub'], value="Non-Billable ODCs")
        col_num = header_start_col

        while col_num <= ws.max_column:
            col_letter = cell_utils.get_column_letter(col_num)
            is_empty = all(ws.cell(row=r, column=col_num).value is None for r in range(7, 9))
            if not is_empty:
                first_row_to_sum = Employee_total_label_row
                last_row_to_sum = Employee_total_label_row + 3

                sum_formula = f"=SUM({col_letter}{first_row_to_sum}:{col_letter}{last_row_to_sum + 1})"
                ws.cell(row=Employee_total_label_row + 5, column=col_num).value = sum_formula
                ws.cell(row=Employee_total_label_row + 5, column=col_num).font = Font(bold=False)
                if col_num >= second_section_start_col:
                    ws.cell(row=Employee_total_label_row + 5, column=col_num).number_format = '$* #,##0_);($* (#,##0);$* 0_);_(@_)'
            col_num += 1  

        non_billable_odcs_row = Employee_total_label_row + 4
        wrap_offset = 12  
        percentage_offset = 6  

        label_column = total_column  
        percentage_column = 22  
        wrap_column = percentage_column - 1
        labels = ["Fringe", "OH", "M&S", "G&A", "Indirects"]

        percentages_df = helper_functions.get_contract_percentages(contract_id)
        percentages = percentages_df.iloc[0].values * 100  

        percentages_row_start = non_billable_odcs_row + percentage_offset  
        fringe_cell = f"{get_column_letter(percentage_column)}{percentages_row_start}"  
        oh_cell = f"{get_column_letter(percentage_column)}{percentages_row_start + 1}"  
        ms_cell = f"{get_column_letter(percentage_column)}{percentages_row_start + 2}"  
        ga_cell = f"{get_column_letter(percentage_column)}{percentages_row_start + 3}"  
        indirects_cell = f"{get_column_letter(percentage_column)}{percentages_row_start + 4}"
       
        # percentages[0] = fringe_cell, percentages[1] = oh_cell, percentages[2] = ms_cell, percentages[3] = ga_cell
        
        # EE Wrap
        ws.cell(row=non_billable_odcs_row + wrap_offset, column=wrap_column, value="EE Wrap")
        ws.cell(row=non_billable_odcs_row + wrap_offset, column=wrap_column + 1, 
                value=f"=ROUND((1*(1+{fringe_cell})*(1+{oh_cell})*(1+{ga_cell})), 2)")
        EE_Wrap = f"{get_column_letter(wrap_column + 1)}{non_billable_odcs_row + wrap_offset}"

        # Sub Wrap
        ws.cell(row=non_billable_odcs_row + wrap_offset + 1, column=wrap_column, value="Sub Wrap")
        ws.cell(row=non_billable_odcs_row + wrap_offset + 1, column=wrap_column + 1, 
                value=f"=ROUND((1*(1+{ms_cell}))+({ms_cell}*{ga_cell}), 2)")
        
        # append formulas 
        for formula in formula_array:
            row = formula['row']
            
            CM_formula = f"=IFERROR(({cell_utils.get_column_letter(col_indices['Billable Rate'])}{row} - ({cell_utils.get_column_letter(col_indices['Direct Rate'])}{row} * (1 + {fringe_cell}))) / {cell_utils.get_column_letter(col_indices['Billable Rate'])}{row}, 0)"
            ws[f"{cell_utils.get_column_letter(col_indices['CM%'])}{row}"].value = CM_formula

            OI_formula = f"=IFERROR(({cell_utils.get_column_letter(col_indices['Billable Rate'])}{row} - ({cell_utils.get_column_letter(col_indices['Direct Rate'])}{row} * {EE_Wrap})) / {cell_utils.get_column_letter(col_indices['Billable Rate'])}{row}, 0)"
            ws[f"{cell_utils.get_column_letter(col_indices['OI%'])}{row}"].value = OI_formula
            
            ws[f"A{row}"].value = formula['countif_formula']

        for i, label in enumerate(labels):
            row = non_billable_odcs_row + percentage_offset + i
            ws.cell(row=row, column=label_column, value=label)
        
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for i, (label, percentage) in enumerate(zip(labels, percentages)):
            row = non_billable_odcs_row + percentage_offset + i
            cell = ws.cell(row=row, column=percentage_column, value=f"{percentage}%")
            cell.fill = yellow_fill
       
       # Init fringe_row
        if 'non_billable_odcs_row' in locals() and 'percentage_offset' in locals():
            fringe_row = non_billable_odcs_row + percentage_offset
        else:
            return
        
        # Fringe formula
        fringe_row = non_billable_odcs_row + percentage_offset  
        for col_num in range(total_col_idx + 1, ws.max_column + 1):  
            col_letter = cell_utils.get_column_letter(col_num)
            employee_labor_value = ws.cell(row=Employee_total_label_row, column=col_num).value

            if employee_labor_value in [None, "",]:
                ws.cell(row=fringe_row, column=col_num).value = ""
            else:
                fringe_formula = f"={col_letter}{Employee_total_label_row}*{fringe_cell}"
                ws.cell(row=fringe_row, column=col_num).value = fringe_formula
                ws.cell(row=fringe_row, column=col_num).font = Font(bold=False)
                if col_num >= second_section_start_col:  
                    ws.cell(row=fringe_row, column=col_num).number_format = '$* #,##0_);($* (#,##0);$* 0_);_(@_)'

        # OH formula
        oh_row = non_billable_odcs_row + percentage_offset + 1 
        for col_num in range(total_col_idx + 1, ws.max_column + 1):  
            col_letter = cell_utils.get_column_letter(col_num)

            fringe_value = ws.cell(row=fringe_row, column=col_num).value
            employee_labor_value = ws.cell(row=Employee_total_label_row, column=col_num).value
            all_values = [fringe_value, employee_labor_value]

            if col_num >= second_section_start_col:  
                ws.cell(row=oh_row, column=col_num).number_format = '$* #,##0_);($* (#,##0);$* 0_);_(@_)'

            if all(val in [None, ""] for val in all_values):
                ws.cell(row=oh_row, column=col_num).value = ""
            else:
                oh_formula = f"=({col_letter}${fringe_row}+{col_letter}${Employee_total_label_row})*{oh_cell}"
                ws.cell(row=oh_row, column=col_num).value = oh_formula

        # M&S formula
        ms_row = non_billable_odcs_row + percentage_offset + 2 
        for col_num in range(total_col_idx + 1, ws.max_column + 1):  
            col_letter = cell_utils.get_column_letter(col_num)

            subcontractor_labor_value = ws.cell(row=subcontractor_total_label_row, column=col_num).value
            direct_software_value = ws.cell(row=Employee_total_label_row + 2, column=col_num).value 
            all_values = [subcontractor_labor_value, direct_software_value]

            if all(val in [None, ""] for val in all_values):
                ws.cell(row=ms_row, column=col_num).value = ""
            else:
                ms_formula = f"=({col_letter}{subcontractor_total_label_row}+{col_letter}{Employee_total_label_row + 2})*{ms_cell}"
                ws.cell(row=ms_row, column=col_num).value = ms_formula
            if col_num >= second_section_start_col:  
                ws.cell(row=ms_row, column=col_num).number_format = '$* #,##0_);($* (#,##0);$* 0_);_(@_)'

        # G&A formula
        ga_row = non_billable_odcs_row + percentage_offset + 3  
        for col_num in range(total_col_idx + 1, ws.max_column + 1):  
            col_letter = cell_utils.get_column_letter(col_num)

            fringe_value = ws.cell(row=fringe_row, column=col_num).value  
            employee_labor_value = ws.cell(row=Employee_total_label_row, column=col_num).value  
            oh_value = ws.cell(row=oh_row, column=col_num).value 
            billable_odc_value = ws.cell(row=Employee_total_label_row + 3, column=col_num).value  
            non_billable_odcs_value = ws.cell(row=Employee_total_label_row + 4, column=col_num).value
            ms_value = ws.cell(row=ms_row, column=col_num).value  
            all_values = [fringe_value, employee_labor_value, oh_value, billable_odc_value, non_billable_odcs_value, ms_value]

            if all(val in [None, ""] for val in all_values):
                ws.cell(row=ga_row, column=col_num).value = ""
            else:
                ga_formula = (
                f"=(SUM({col_letter}{fringe_row},{col_letter}{Employee_total_label_row},"
                f"{col_letter}{oh_row},{col_letter}{Employee_total_label_row + 3},"
                f"{col_letter}{Employee_total_label_row + 4})*{ga_cell}) + ({col_letter}{ms_row}*{ga_cell})"
                )
                ws.cell(row=ga_row, column=col_num).value = ga_formula
            if col_num >= second_section_start_col:  
                ws.cell(row=ga_row, column=col_num).number_format = '$* #,##0_);($* (#,##0);$* 0_);_(@_)'

        # Indirects formula
        indirects_row = non_billable_odcs_row + percentage_offset + 4 
        for col_num in range(total_col_idx + 1, ws.max_column + 1):  
            col_letter = cell_utils.get_column_letter(col_num)

            fringe_value = ws.cell(row=fringe_row, column=col_num).value  
            oh_value = ws.cell(row=oh_row, column=col_num).value  
            ms_value = ws.cell(row=ms_row, column=col_num).value  
            ga_value = ws.cell(row=ga_row, column=col_num).value 
            all_values = [fringe_value, oh_value, ms_value, ga_value]

            if all(val in [None, ""] for val in all_values):
                ws.cell(row=indirects_row, column=col_num).value = ""
            else:
                indirects_formula = f"=SUM({col_letter}{fringe_row}:{col_letter}{ga_row})"
                ws.cell(row=indirects_row, column=col_num).value = indirects_formula

            if col_num >= second_section_start_col:  
                ws.cell(row=indirects_row, column=col_num).number_format = '$* #,##0_);($* (#,##0);$* 0_);_(@_)'

        ws.freeze_panes = 'J9'

    except Exception as e:
        print(f"Error creating direct costs spreadsheet: {e}")