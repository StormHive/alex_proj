from openpyxl.styles import Font, PatternFill
from utils.helper_functions import list_contract_months, set_cell, set_font_for_sheet, add_footer_rows, holidays_formula
from datetime import datetime
import pandas as pd
import utils.helper_functions as helper_functions

# Create HRS spreadsheet
def create_HRS_Spreadsheet(wb, headers_data, data, lookup_month_info, last_month_dt, PeriodNumber, engine):
    Sheet_title = f"OY{PeriodNumber} HRS" if PeriodNumber >= 1 else "Base HRS"
    ws = wb.create_sheet(title=Sheet_title)
    set_font_for_sheet(ws)
    
    bold_font = Font(bold=True)
    salmon_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")

    headers = [
        ("A1", "Prime Contract #", headers_data['prime contract #'].iloc[0]),
        ("A2", "Task #", headers_data['task #'].iloc[0]),
        ("A3", "Contract Name", headers_data['contract name'].iloc[0]),
        ("A4", "POP", headers_data['pop'].iloc[0]),
        ("A5", "DM", headers_data['dm'].iloc[0]),
        ("A6", "PM", headers_data['pm'].iloc[0])
    ]
    for cell, header, value in headers:
        set_cell(ws, cell, header, bold=True)
        set_cell(ws, cell.replace('A', 'B'), value, bold=True)

    set_cell(ws, 'B8', "Work % less PTO Est", bold=True)
    set_cell(ws, 'A10', "Labor Forecast:", bold=True)
    set_cell(ws, 'A11', "Labor Category", bold=True)
    set_cell(ws, 'B11', "Name", bold=True)
    set_cell(ws, 'C11', "Avail Hrs", bold=True)
    set_cell(ws, 'D10', "Work Hrs", bold=True)
    set_cell(ws, 'D11', "Percentage", bold=True)

    # POP Dates
    pop_start_date = pd.Timestamp(headers_data['pop'].iloc[0].split(' - ')[0])
    pop_end_date = pd.Timestamp(headers_data['pop'].iloc[0].split(' - ')[1])
    generated_months = list_contract_months(pop_start_date.strftime("%Y-%m-%d"), pop_end_date.strftime("%Y-%m-%d"))
    total_col_index = len(generated_months) + 5
    total_col_letter = chr(ord('A') + total_col_index - 1)

    for row in range(8, 12):
        for col in range(1, total_col_index + 1): 
            ws.cell(row=row, column=col).fill = salmon_fill

    # Calculate workdays each month 
    generate_formula, holiday_list = holidays_formula(engine, pop_start_date, pop_end_date)
    for i, month in enumerate(generated_months, start=5):
        month_col_letter = chr(ord('A') + i - 1)
        set_cell(ws, f'{month_col_letter}9', month)
        ws.cell(row=9, column=i).number_format = 'mmm-yy'
        month_datetime = datetime.strptime(month, '%b-%y')
        
        start_day = (
            pop_start_date.day if month_datetime.month == pop_start_date.month and month_datetime.year == pop_start_date.year
            else 1
        )
        end_day = (
            pop_end_date.day if month_datetime.month == pop_end_date.month and month_datetime.year == pop_end_date.year
            else pd.Timestamp(year=month_datetime.year, month=month_datetime.month, day=1).days_in_month
        )

        formula = generate_formula(month_datetime.year, month_datetime.month, start_day, end_day)
        set_cell(ws, f'{month_col_letter}10', formula)
        set_cell(ws, f'{month_col_letter}7', f'=8 * {month_col_letter}10')

    set_cell(ws, f'{total_col_letter}9', "Total", bold=True)
    set_cell(ws, f'{total_col_letter}10', f"=SUM(E10:{chr(ord('A') + len(generated_months) + 3)}10)", bold=True)

    start_row = 12
    start_row = create_HRS_employee_rows(ws, data, lookup_month_info, start_row, generated_months, last_month_dt, headers_data, engine)
    add_footer_rows(ws, start_row, generated_months)
    return wb

# Create HRS employee rows
def create_HRS_employee_rows(ws, data, lookup_month_info, start_row, generated_months, last_month_dt, headers_data, engine):
    pop_start_date = pd.Timestamp(headers_data['pop'].iloc[0].split(' - ')[0])
    pop_end_date = pd.Timestamp(headers_data['pop'].iloc[0].split(' - ')[1])
    
    grouped_data = data.groupby('JobTitle')
    for job_name, group in grouped_data:
        set_cell(ws, f'A{start_row}', job_name.split(":")[0], bold=True)
        for col in range(1, 5 + len(generated_months) + 1):
            ws.cell(row=start_row, column=col).fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        start_row += 1 
        current_employee = None

        for _, row in group.iterrows():
            if row['EmployeeName'] != current_employee:
                current_employee = row['EmployeeName']
                set_cell(ws, f'A{start_row}', row['LaborCategoryName'])
                if row['EmployeeName'] == 'TBD':
                    set_cell(ws, f'B{start_row}', f"TBD - {row['Company']} - {row['PreviousEmployeeLastName']} Replacement")
                else:
                    set_cell(ws, f'B{start_row}', row['EmployeeName'])
                    set_cell(ws, f'C{start_row}', row['AvailableHours'])
                    set_cell(ws, f'D{start_row}', row['WorkHoursPercentage'], bold=False)
                    ws[f'D{start_row}'].number_format = '0%'

                for month_col, month in enumerate(generated_months, start=5):
                    month_col_letter = chr(ord('A') + month_col - 1)
                    month_datetime = datetime.strptime(month, '%b-%y')

                    # Vacation factor percentage
                    vacation_factor_percentage = lookup_month_info[
                        (lookup_month_info['Month'] == month_datetime.month) &
                        (lookup_month_info['Year'] == month_datetime.year)
                    ]['VacationFactor'].values
                    if vacation_factor_percentage:
                        set_cell(ws, f"{month_col_letter}8", vacation_factor_percentage[0], bold=False)
                        ws[f"{month_col_letter}8"].number_format = '0.00%'
                   
                    if month_datetime <= last_month_dt: 
                        hours_worked = data[
                            (data['employee_id'] == row.employee_id) &
                            (data['date'].dt.month == month_datetime.month) &
                            (data['date'].dt.year == month_datetime.year)
                        ]['HoursToDisplayInCell'].sum()
                        set_cell(ws, f'{month_col_letter}{start_row}', hours_worked)
                        ws[f'{month_col_letter}{start_row}'].number_format = '0.0'

                    else:
                        hours_available_override = data[
                            (data['employee_id'] == row.employee_id) &
                            (data['date'].dt.month == month_datetime.month) &
                            (data['date'].dt.year == month_datetime.year)
                        ]['HoursToDisplayInCell'].values[0]
                        if pd.isnull(hours_available_override):
                            final_value = f'=({month_col_letter}$10*8)*$D{start_row}*${month_col_letter}$8'
                        else:
                            final_value = hours_available_override
                        set_cell(ws, f'{month_col_letter}{start_row}', final_value, fill=PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))
                        ws[f'{month_col_letter}{start_row}'].number_format = '#,##0.0_);(#,##0.0)'

                start_row += 1

        # Totals for each row
        total_col_index = len(generated_months) + 5
        total_col_letter = chr(ord('A') + total_col_index - 1)
        for idx in range(12, start_row + 1):
            sum_range = f"E{idx}:{chr(ord('A') + len(generated_months) + 3)}{idx}"
            all_empty = all(ws.cell(row=idx, column=col).value in [None, ""] for col in range(5, len(generated_months) + 5))
            if not all_empty: 
                set_cell(ws, f'{total_col_letter}{idx}', f"=ROUND(SUM({sum_range}), 1)", bold=True)
                ws[f'{total_col_letter}{idx}'].number_format = '0.0'

        # Total for each column
        footer_row = start_row + 1
        for i, month in enumerate(generated_months, start=5):
            month_col_letter = chr(ord('A') + i - 1)
            set_cell(ws, f'{month_col_letter}{footer_row}', f"=SUM({month_col_letter}13:{month_col_letter}{start_row})")
            set_cell(ws, f'{total_col_letter}{footer_row}', f"=SUM({total_col_letter}13:{total_col_letter}{start_row})", bold=True)

        # Count HRS sheet values
        count_row = footer_row + 2
        for i, month in enumerate(generated_months, start=5):
            month_col_letter = chr(ord('A') + i - 1)
            set_cell(ws, f'{month_col_letter}{count_row}', f"=COUNTA({month_col_letter}13:{month_col_letter}{start_row})-COUNTIF({month_col_letter}13:{month_col_letter}{start_row},0)", bold=False)
            ws[f'{month_col_letter}{count_row}'].number_format = '0.0'

    return start_row
